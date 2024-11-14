from itertools import product,combinations
from itertools import permutations
import openpyxl
def read_truth_tables(filename):
    global wb

    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    first_table = []
    second_table = []
    switch_table = False
    for row in ws.iter_rows(values_only=True):

        if not(switch_table) and row[0]==2:  # Пустая строка, переключаем таблицы
            switch_table = True
            continue
        if row[0]!= 1 and row[0]!= 0 and row[0]!= None:
            continue

        if not switch_table :

            first_table.append([row[cell] for cell in range(4) if row[6]])
        else:
            second_table.append([row[cell] for cell in range(4) ])

    return first_table, second_table
def place_in_orded(filename,asd):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    sheet = wb.active

    # Вставляем 4 буквы в 8 столбец (нумерация начинается с 1)
    col = 8  # Столбец H
    data = [asd]  # Данные для вставки

    for row, entry in enumerate(data, start=1):
        sheet.cell(row=row, column=col, value=entry)

    # Сохраняем изменения в файл
    wb.save(str(filename))
# filename = "C:\\Users\\s09t0\\Desktop\\ексель.xlsx"
filename = input('введите путь к xls файлу: ')
try:

    filename = filename.replace("\\", "\\\\")
    matrix, uncomplited = read_truth_tables(filename)


except:

    print('некоректный путь')
    k=input("нажмите enter чтобы закрыть")
def find_otv(matrix,uncomplited):
    def get_matrix_permutations(matrix):

        matrix_permutations = []
        for perm in permutations(matrix):
            matrix_permutations.append(list(perm))
        return matrix_permutations


    def columns_matrix(matrix):

        matrix1=get_matrix_permutations(matrix)

        completed_vertical = []
        otv=[]
        for matrix in matrix1:
            completed_vertical = []
            for i in range(4):  # переворачиваем в вертикальную матрицу полную
                completed_vertical.append([])
                for j in range(3):
                    completed_vertical[-1].append(matrix[j][i])
            abc='xyzw'
            for i in range(len(completed_vertical)):
                completed_vertical[i].insert(0,abc[i])
            otv.append(completed_vertical)
        return completed_vertical
    def permutations_without_module(lst): # все перестановки полной матрицы
        if len(lst) == 0:
            return [[]]
        permutations = []
        for i in range(len(lst)):
            current = lst[i]
            remaining = lst[:i] + lst[i + 1:]
            for p in permutations_without_module(remaining):
                permutations.append([current] + p)
        return permutations

    def generate_combinations(matrix): #для неполной меняем  все None на 1 или 0 (все комбинации)
        none_positions = [(i, j) for i in range(len(matrix)) for j in range(len(matrix[0])) if matrix[i][j] is None]
        replacements = product([0, 1], repeat=len(none_positions))
        all_combinations = []
        for replace_values in replacements:
            temp_matrix = [row[:] for row in matrix]
            for (i, j), replacement in zip(none_positions, replace_values):
                temp_matrix[i][j] = replacement
            all_combinations.append(temp_matrix)
        return all_combinations


    columns = columns_matrix(matrix) #полная матрица


    # Генерация всех возможных перестановок
    all_permutations_completed = permutations_without_module(columns)
    # Вывод всех перестановок


    # с неполной матрицей

    uncomplited_vertical=[]
    for i in range(4): #переворачиваем в вертикальную матрицу
        uncomplited_vertical.append([])
        for j in range(3):
            uncomplited_vertical[-1].append(uncomplited[j][i])
    combinations_uncomplited=generate_combinations(uncomplited_vertical)

    def perebor(all_permutations_completed,combinations_uncomplited):
        for x in all_permutations_completed:
            now=[z[1:] for z in x]
            for y in combinations_uncomplited:
                if now==y:
                    return x
    return[x[0] for x in perebor(all_permutations_completed,combinations_uncomplited)]


def get_array_combinations(array):
    array_combinations = []
    for combination in combinations(array, 3):
        array_combinations.append(list(combination))
    return array_combinations

matrix=get_array_combinations(matrix)

for x in matrix:
    try:
        now=find_otv(x,uncomplited)
    except:
        print('что то пошло не так возможно проблема с таблицей.')
        break
    if now!= None:
          place_in_orded(filename,''.join(now))
          print(now, '  столбцы успешно рассчитаны ')
# def find_one():
# print(all_permutations_completed)
# print(combinations_uncomplited)


k=input("нажмите enter чтобы закрыть")